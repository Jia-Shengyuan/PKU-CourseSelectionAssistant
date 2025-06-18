import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from api.models.course import Course
from typing import List, Dict, Tuple
import re
from datetime import datetime

# 课程颜色配置 - 与前端保持一致
colorPalette = [
    '#FFEBEE', '#E3F2FD', '#E8F5E9', '#FFFDE7', '#F3E5F5', '#FBE9E7', '#E0F2F1', '#FFF3E0', '#F9FBE7', '#ECEFF1',
    '#FFCDD2', '#BBDEFB', '#C8E6C9', '#FFF9C4', '#D1C4E9', '#FFCCBC', '#B2DFDB', '#FFE0B2', '#F0F4C3', '#CFD8DC'
]

def parse_course_time(time_str: str) -> List[Dict]:
    """
    解析课程时间字符串
    例：'all,1,3-4;all,3,1-2;' -> [{'week_type': 'all', 'day': 1, 'start': 3, 'end': 4}, ...]
    """
    if not time_str:
        return []
    
    result = []
    segments = [seg.strip() for seg in time_str.split(';') if seg.strip()]
    
    for segment in segments:
        parts = segment.split(',')
        if len(parts) < 3:
            continue
            
        week_type = parts[0]  # all, odd, even
        try:
            day = int(parts[1])  # 1=周一, 2=周二, ...
            time_range = parts[2]
            
            if '-' in time_range:
                start, end = map(int, time_range.split('-'))
            else:
                start = end = int(time_range)
                
            result.append({
                'week_type': week_type,
                'day': day,
                'start': start,
                'end': end
            })
        except (ValueError, IndexError):
            continue
    
    return result

def get_week_type_text(week_type: str) -> str:
    """将weekType转换为中文显示"""
    mapping = {
        'all': '每周',
        'odd': '单周', 
        'even': '双周'
    }
    return mapping.get(week_type, '每周')

def format_teacher_name(teacher_name: str) -> str:
    """格式化教师姓名，移除拼音缩写部分"""
    if not teacher_name:
        return ''
    
    names = teacher_name.split(',')
    formatted_names = []
    
    for name in names:
        name = name.strip()
        # 查找 '/' 字符的位置
        slash_index = name.find('/')
        if slash_index != -1:
            # 找到 '/' 后，移除 '/' 及其后面的连续英文字母
            before_slash = name[:slash_index]
            after_slash = name[slash_index + 1:]
            # 移除连续的英文字母，保留其他字符（如括号和中文）
            cleaned_after_slash = re.sub(r'^[a-zA-Z]+', '', after_slash)
            formatted_names.append(before_slash + cleaned_after_slash)
        else:
            formatted_names.append(name)
    
    return '，'.join(formatted_names)

def create_course_color_map(timetable: List[List[Course]]) -> Dict[str, str]:
    """为课程分配颜色"""
    course_names = set()
    for plan in timetable:
        for course in plan:
            if course.name:
                course_names.add(course.name)
    
    color_map = {}
    for i, course_name in enumerate(sorted(course_names)):
        color_map[course_name] = colorPalette[i % len(colorPalette)]
    
    return color_map

def hex_to_rgb(hex_color: str) -> Tuple[int, int, int]:
    """将十六进制颜色转换为RGB"""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def create_timetable_grid(courses: List[Course], color_map: Dict[str, str]):
    """创建课表网格数据"""
    # 初始化12x7的网格 (12节课 x 7天)
    grid = [[None for _ in range(7)] for _ in range(12)]
    
    for course in courses:
        if not course.time:
            continue
            
        time_slots = parse_course_time(course.time)
        bg_color = color_map.get(course.name, '#FFFFFF')
        
        for slot in time_slots:
            day_idx = slot['day'] - 1  # 转换为0基索引
            start_period = slot['start'] - 1
            end_period = slot['end'] - 1
            
            if 0 <= day_idx < 7 and 0 <= start_period < 12 and 0 <= end_period < 12:
                for period in range(start_period, end_period + 1):
                    if period == start_period:
                        # 主单元格
                        grid[period][day_idx] = {
                            'course': course,
                            'week_type': get_week_type_text(slot['week_type']),
                            'bg_color': bg_color,
                            'rowspan': end_period - start_period + 1,
                            'is_merged': False
                        }
                    else:
                        # 被合并的单元格
                        grid[period][day_idx] = {
                            'course': None,
                            'week_type': '',
                            'bg_color': bg_color,
                            'rowspan': 0,
                            'is_merged': True
                        }
    
    return grid

def export_to_excel(timetable: List[List[Course]], file_name='课表.xlsx'):
    """导出课表到Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 为课程分配颜色
    color_map = create_course_color_map(timetable)
    
    # 为每个方案创建一个工作表
    for i, courses in enumerate(timetable):
        ws = wb.create_sheet(title=f"方案{i+1}")
        
        # 创建课表网格
        grid = create_timetable_grid(courses, color_map)
        
        # 设置表头
        headers = ['节次', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(name='微软雅黑', size=11, bold=True)
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 填充课表数据
        for period in range(12):
            row = period + 2  # 从第2行开始（第1行是表头）
            
            # 设置时间列
            time_cell = ws.cell(row=row, column=1)
            time_cell.value = f"第{period+1}节"
            time_cell.font = Font(name='微软雅黑', size=10, bold=True)
            time_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            time_cell.alignment = Alignment(horizontal='center', vertical='center')
            time_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置课程列
            for day in range(7):
                col = day + 2  # 从第2列开始（第1列是时间）
                cell = ws.cell(row=row, column=col)
                cell_data = grid[period][day]
                
                if cell_data and not cell_data['is_merged'] and cell_data['course']:
                    course = cell_data['course']
                    
                    # 设置单元格内容
                    teacher_name = format_teacher_name(course.teacher) if course.teacher else ''
                    location = course.location if course.location else ''
                    week_location = f"{cell_data['week_type']} {location}".strip()
                    
                    cell_text = course.name
                    if teacher_name:
                        cell_text += f"\n{teacher_name}"
                    if week_location:
                        cell_text += f"\n{week_location}"
                    
                    cell.value = cell_text
                    
                    # 设置字体样式
                    cell.font = Font(name='微软雅黑', size=10)
                    
                    # 设置背景颜色
                    rgb = hex_to_rgb(cell_data['bg_color'])
                    cell.fill = PatternFill(
                        start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                        end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                        fill_type='solid'
                    )
                    
                    # 合并单元格
                    if cell_data['rowspan'] > 1:
                        ws.merge_cells(
                            start_row=row,
                            start_column=col,
                            end_row=row + cell_data['rowspan'] - 1,
                            end_column=col
                        )
                
                elif cell_data and cell_data['is_merged']:
                    # 被合并的单元格，只设置背景色
                    rgb = hex_to_rgb(cell_data['bg_color'])
                    cell.fill = PatternFill(
                        start_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                        end_color=f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                        fill_type='solid'
                    )
                else:
                    # 空课时
                    cell.value = ""
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # 设置单元格样式
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8  # 节次列
        for col in range(2, 9):  # 周一到周日
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # 设置行高
        ws.row_dimensions[1].height = 25  # 表头行高
        for row in range(2, 14):  # 数据行
            ws.row_dimensions[row].height = 65
    
    # 生成文件名
    if not file_name.endswith('.xlsx'):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"课表方案_{timestamp}.xlsx"
    
    # 保存文件
    wb.save(file_name)
    return file_name